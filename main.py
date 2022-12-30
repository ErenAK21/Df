import openpyxl
import os

# Erstelle eine neue Excel-Datei
wb = openpyxl.Workbook()

# Wähle das erste Arbeitsblatt aus
ws = wb.active
row_number = 2

# Setze Überschriften für die Spalten
ws.cell(row=1, column=1).value = "Fahrzeug"
ws.cell(row=1, column=2).value = "Frist"
ws.cell(row=1, column=3).value = "Datum"
ws.cell(row=1, column=4).value = "Werkstatt"
ws.cell(row=1, column=5).value = "ASNR"

# Gehe durch alle Dateien im Ordner
for filename in os.listdir(r"C:\Users\eaksu\OneDrive - National Express Rail GmbH\Desktop\Arbeitsscheine haufen"):
    # Trenne den Dateinamen anhand des letzten Unterstrichs
    parts = filename.rsplit("_", 1)

    # Wenn der Dateiname den erwarteten Aufbau hat, füge die Informationen der Tabelle hinzu
    if len(parts) == 2:
        # Trenne den Teil vor dem letzten Unterstrich noch einmal anhand des Unterstrichs
        subparts = parts[0].split("_")
        if len(subparts) == 5:
            ws.cell(row=row_number, column=1).value = subparts[0]
            ws.cell(row=row_number, column=2).value = subparts[1]
            ws.cell(row=row_number, column=3).value = subparts[2] + "-" + subparts[3]
            ws.cell(row=row_number, column=4).value = subparts[4]  # Entferne die Dateiendung ".pdf"
            ws.cell(row=row_number, column=5).value = parts[1][:-4]
            row_number += 1

# Speichere die Excel-Datei
wb.save("meine_tabelle.xlsx")